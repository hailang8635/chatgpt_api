# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
from openpyxl import load_workbook

# Create score difference table data
score_diff_data = {
    'Score_Range': ['0-10', '11-40', '41-70', '71-100', '101-130', '131-160', 
                   '161-190', '191-220', '221-250', '251+'],
    'Higher_Score': [9, 8, 7, 6, 5, 4, 3, 2, 1, 0],
    'Lower_Score': [9, 12, 16, 20, 25, 30, 35, 40, 45, 50]
}

# Create match results data
players = ['Sun Jiaoqin', 'Huang Yuchen', 'Zhang Mengxin', 'Kong Zihan', 'Xu Yiyang']
ratings = [1000, 1000, 1036, 966, 974]

# Create empty DataFrame for match results with initial format
match_results = pd.DataFrame('', index=players, columns=players)

# Save to Excel
with pd.ExcelWriter('tournament_scores.xlsx', engine='openpyxl') as writer:
    # Save score difference table
    score_diff_df = pd.DataFrame(score_diff_data)
    score_diff_df.to_excel(writer, sheet_name='Score Rules', index=False)
    
    # Create match results table
    result_df = pd.DataFrame(index=players)
    result_df['Rating'] = ratings
    result_df = pd.concat([result_df, match_results], axis=1)
    result_df['Total_Score'] = 0
    result_df['Rank'] = ''
    
    # Save to Excel
    result_df.to_excel(writer, sheet_name='Match Results')

# Load workbook for formatting
wb = load_workbook('tournament_scores.xlsx')

# Format Score Rules sheet
ws_rules = wb['Score Rules']
for col in ['A', 'B', 'C']:
    ws_rules.column_dimensions[col].width = 15

# Format Match Results sheet
ws_match = wb['Match Results']
ws_match.insert_rows(1)
ws_match.cell(row=1, column=1, value="B-1 Group  Table 2  Qualifiers: 4  Check-in: 7:45-7:55")

# Set column widths
ws_match.column_dimensions['A'].width = 15
for col in range(ord('B'), ord('B') + len(players) + 3):
    ws_match.column_dimensions[chr(col)].width = 12

# Save formatted workbook
wb.save('tournament_scores.xlsx')

print("Tournament score template created: tournament_scores.xlsx")
print("Please fill in match results in format W:L (e.g. 2:1)") 